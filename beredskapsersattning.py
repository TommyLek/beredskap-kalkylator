from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Skapa arbetsbok
wb = Workbook()
ws = wb.active
ws.title = "Beredskapsersättning"

# Stilar
header_font = Font(bold=True, size=12)
title_font = Font(bold=True, size=14)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, color="FFFFFF")

# Titel
ws['A1'] = "Beredskapsersättning - Beredskapstjänst II"
ws['A1'].font = title_font
ws.merge_cells('A1:E1')

ws['A2'] = "Handelns tjänstemannaavtal (1 maj 2025 - 30 april 2027)"
ws.merge_cells('A2:E2')

# Inmatningssektion
ws['A4'] = "INMATNING"
ws['A4'].font = header_font

ws['A5'] = "Månadslön:"
ws['B5'] = 40000
ws['B5'].number_format = '#,##0 "kr"'
ws['B5'].font = Font(bold=True)

# Timmar per ersättningsnivå
ws['A7'] = "TIMMAR PER VECKA (normalvecka utan helgdag)"
ws['A7'].font = header_font
ws.merge_cells('A7:E7')

headers = ["Ersättningsnivå", "Period", "Timmar", "Divisor", "Ersättning/vecka"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=8, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Data
data = [
    ("Grundersättning", "Mån-fre kvällar/nätter + mån morgon", 69, 1400),
    ("Fredagskväll/natt", "Fre 18:00 - Lör 07:00", 13, 1000),
    ("Helgersättning", "Lör 07:00 - Sön 24:00", 41, 700),
]

for row_idx, (level, period, hours, divisor) in enumerate(data, 9):
    ws.cell(row=row_idx, column=1, value=level).border = thin_border
    ws.cell(row=row_idx, column=2, value=period).border = thin_border
    cell_hours = ws.cell(row=row_idx, column=3, value=hours)
    cell_hours.border = thin_border
    cell_hours.alignment = Alignment(horizontal='center')
    cell_div = ws.cell(row=row_idx, column=4, value=divisor)
    cell_div.border = thin_border
    cell_div.alignment = Alignment(horizontal='center')
    # Formel för ersättning
    cell_ers = ws.cell(row=row_idx, column=5)
    cell_ers.value = f"=C{row_idx}*$B$5/D{row_idx}"
    cell_ers.number_format = '#,##0.00 "kr"'
    cell_ers.border = thin_border

# Summa
ws['A12'] = "TOTALT"
ws['A12'].font = header_font
ws['C12'] = "=SUM(C9:C11)"
ws['C12'].font = header_font
ws['C12'].alignment = Alignment(horizontal='center')
ws['E12'] = "=SUM(E9:E11)"
ws['E12'].number_format = '#,##0.00 "kr"'
ws['E12'].font = header_font

for col in range(1, 6):
    ws.cell(row=12, column=col).border = thin_border

# Procent av månadslön
ws['A14'] = "Andel av månadslön:"
ws['B14'] = "=E12/B5"
ws['B14'].number_format = '0.00%'
ws['B14'].font = Font(bold=True)

# Detaljerad uppdelning
ws['A16'] = "DETALJERAD TIDFÖRDELNING"
ws['A16'].font = header_font

detail_headers = ["Dag", "Period", "Timmar", "Ersättningsnivå"]
for col, header in enumerate(detail_headers, 1):
    cell = ws.cell(row=17, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border

details = [
    ("Måndag", "18:00-24:00", 6, "Grund (1/1400)"),
    ("Tisdag", "00:00-09:00", 9, "Grund (1/1400)"),
    ("Tisdag", "18:00-24:00", 6, "Grund (1/1400)"),
    ("Onsdag", "00:00-09:00", 9, "Grund (1/1400)"),
    ("Onsdag", "18:00-24:00", 6, "Grund (1/1400)"),
    ("Torsdag", "00:00-09:00", 9, "Grund (1/1400)"),
    ("Torsdag", "18:00-24:00", 6, "Grund (1/1400)"),
    ("Fredag", "00:00-09:00", 9, "Grund (1/1400)"),
    ("Fredag", "18:00-24:00", 6, "Fredagskväll (1/1000)"),
    ("Lördag", "00:00-07:00", 7, "Fredagskväll (1/1000)"),
    ("Lördag", "07:00-24:00", 17, "Helg (1/700)"),
    ("Söndag", "00:00-24:00", 24, "Helg (1/700)"),
    ("Måndag", "00:00-09:00", 9, "Grund (1/1400)"),
]

for row_idx, (dag, period, timmar, nivå) in enumerate(details, 18):
    ws.cell(row=row_idx, column=1, value=dag).border = thin_border
    ws.cell(row=row_idx, column=2, value=period).border = thin_border
    cell_t = ws.cell(row=row_idx, column=3, value=timmar)
    cell_t.border = thin_border
    cell_t.alignment = Alignment(horizontal='center')
    ws.cell(row=row_idx, column=4, value=nivå).border = thin_border

# Räkneexempel
ws['A33'] = "RÄKNEEXEMPEL"
ws['A33'].font = header_font

example_headers = ["Månadslön", "Ersättning/vecka"]
for col, header in enumerate(example_headers, 1):
    cell = ws.cell(row=34, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border

salaries = [30000, 35000, 40000, 45000, 50000]
for row_idx, salary in enumerate(salaries, 35):
    cell_sal = ws.cell(row=row_idx, column=1, value=salary)
    cell_sal.number_format = '#,##0 "kr"'
    cell_sal.border = thin_border
    cell_ers = ws.cell(row=row_idx, column=2)
    cell_ers.value = f"={salary}*(69/1400+13/1000+41/700)"
    cell_ers.number_format = '#,##0.00 "kr"'
    cell_ers.border = thin_border

# Kolumnbredder
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 18

# Spara
filepath = "C:/Claude/Beredskap/Beredskapsersattning_kalkylator.xlsx"
wb.save(filepath)
print(f"Excel-fil skapad: {filepath}")
